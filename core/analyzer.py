import os
import math
import pandas as pd
import numpy as np
from openpyxl.styles import PatternFill, Font, Alignment
from datetime import datetime

from core.database import DatabaseManager


def _norm_cdf(x):
    return 0.5 * (1 + math.erf(x / math.sqrt(2)))


def _norm_ppf(p):
    a = [-3.969683028665376e+01, 2.209460984245205e+02,
         -2.759285104469687e+02, 1.383577518672690e+02,
         -3.066479806614716e+01, 2.506628277459239e+00]
    b = [-5.447609879822406e+01, 1.615858368580409e+02,
         -1.556989798598866e+02, 6.680131188771972e+01,
         -1.328068155288572e+01]
    c = [-7.784894002430293e-03, -3.223964580411365e-01,
         -2.400758277161838e+00, -2.549732539343734e+00,
         4.374664141464968e+00, 2.938163982698783e+00]
    d = [7.784695709041462e-03, 3.224671290700398e-01,
         2.445134137142996e+00, 3.754408661907416e+00]
    p_low, p_high = 0.02425, 1 - 0.02425
    if p < p_low:
        q = math.sqrt(-2 * math.log(p))
        return (((((c[0]*q+c[1])*q+c[2])*q+c[3])*q+c[4])*q+c[5]) / \
               ((((d[0]*q+d[1])*q+d[2])*q+d[3])*q+1)
    elif p <= p_high:
        q = p - 0.5
        r = q * q
        return (((((a[0]*r+a[1])*r+a[2])*r+a[3])*r+a[4])*r+a[5])*q / \
               (((((b[0]*r+b[1])*r+b[2])*r+b[3])*r+b[4])*r+1)
    else:
        q = math.sqrt(-2 * math.log(1 - p))
        return -(((((c[0]*q+c[1])*q+c[2])*q+c[3])*q+c[4])*q+c[5]) / \
                ((((d[0]*q+d[1])*q+d[2])*q+d[3])*q+1)


def _t_cdf(t, df):
    x = df / (df + t * t)
    if t >= 0:
        return 1 - 0.5 * _beta_inc(df / 2, 0.5, x)
    else:
        return 0.5 * _beta_inc(df / 2, 0.5, x)


def _beta_inc(a, b, x):
    if x < 0 or x > 1:
        return 0
    if x == 0 or x == 1:
        return x
    lbeta = math.lgamma(a + b) - math.lgamma(a) - math.lgamma(b) + a * math.log(x) + b * math.log(1 - x)
    if x < (a + 1) / (a + b + 2):
        return math.exp(lbeta) * _beta_cf(a, b, x) / a
    else:
        return 1 - math.exp(lbeta) * _beta_cf(b, a, 1 - x) / b


def _beta_cf(a, b, x):
    max_iter, eps = 200, 1e-14
    qab, qap, qam = a + b, a + 1, a - 1
    c, d = 1.0, 1 - qab * x / qap
    if abs(d) < 1e-30:
        d = 1e-30
    d = 1.0 / d
    h = d
    for m in range(1, max_iter + 1):
        m2 = 2 * m
        aa = m * (b - m) * x / ((qam + m2) * (a + m2))
        d = 1 + aa * d
        if abs(d) < 1e-30:
            d = 1e-30
        c = 1 + aa / c
        if abs(c) < 1e-30:
            c = 1e-30
        d = 1.0 / d
        h *= d * c
        aa = -(a + m) * (qab + m) * x / ((a + m2) * (qap + m2))
        d = 1 + aa * d
        if abs(d) < 1e-30:
            d = 1e-30
        c = 1 + aa / c
        if abs(c) < 1e-30:
            c = 1e-30
        d = 1.0 / d
        delta = d * c
        h *= delta
        if abs(delta - 1) < eps:
            break
    return h


def _binom_cdf(k, n, p):
    if p <= 0:
        return 1.0 if k >= n else 0.0
    if p >= 1:
        return 1.0
    total = 0.0
    for i in range(k + 1):
        total += math.comb(n, i) * (p ** i) * ((1 - p) ** (n - i))
    return min(total, 1.0)


def _sem(series):
    n = len(series)
    if n <= 1:
        return 0
    return series.std(ddof=1) / math.sqrt(n)


def _t_interval(conf, df, loc, scale):
    alpha = 1 - conf
    t_val = _norm_ppf(1 - alpha / 2)
    if df > 1:
        t_val *= (1 + (t_val ** 2 + 1) / (4 * df))
    return loc - t_val * scale, loc + t_val * scale


class MarketAnalyzer:
    def __init__(self, db: DatabaseManager):
        self.db = db
        self.iqr_multiplier = self._load_iqr_config()

    def _load_iqr_config(self):
        try:
            cfg_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "analyzer_config.json")
            if os.path.exists(cfg_path):
                import json
                with open(cfg_path, "r") as f:
                    cfg = json.load(f)
                return float(cfg.get("iqr_multiplier", 1.0))
        except:
            pass
        return 1.0

    def get_item_stats(self, item_name, item_lvl, time_limit_minutes=None, start_date=None, end_date=None, server=None):
        try:
            if isinstance(item_name, str):
                item_name = item_name.strip()
            if isinstance(item_lvl, str):
                item_lvl = item_lvl.strip()

            if item_lvl in ["+0", "0", "", None]:
                lvl_condition = "(item_lvl = '' OR item_lvl IS NULL OR item_lvl = '+0' OR item_lvl = '0')"
                params = [item_name]
            else:
                lvl_condition = "item_lvl = ?"
                params = [item_name, item_lvl]

            query = f"SELECT type, price FROM prices WHERE item_name = ? AND {lvl_condition}"

            if server:
                query += " AND server LIKE ?"
                params.append(f"%{server}%")

            if start_date:
                query += " AND timestamp >= ?"
                params.append(start_date[:10])
            elif time_limit_minutes:
                from datetime import timedelta
                cutoff_time = (datetime.utcnow() - timedelta(minutes=int(time_limit_minutes))).strftime('%Y-%m-%d')
                query += " AND timestamp >= ?"
                params.append(cutoff_time)

            if end_date:
                query += " AND timestamp <= ?"
                params.append(end_date[:10])

            with self.db.get_connection() as conn:
                df = pd.read_sql_query(query, conn, params=tuple(params))

            if df.empty:
                return None

            sell_prices = df[df['type'].str.capitalize() == 'Sell']['price']
            buy_prices = df[df['type'].str.capitalize() == 'Buy']['price']

            if len(sell_prices) < 3 and len(buy_prices) < 3:
                return None

            sell_mode = sell_prices.mode().iloc[0] if not sell_prices.empty else 0
            buy_mode = buy_prices.mode().iloc[0] if not buy_prices.empty else 0

            relational_stats = {"covariance": 0, "correlation": 0}

            if not sell_prices.empty and not buy_prices.empty:
                s_clean = self.calculate_filtered_series(sell_prices)
                b_clean = self.calculate_filtered_series(buy_prices)
                min_len = min(len(s_clean), len(b_clean))

                if min_len > 1:
                    s_series = s_clean.sort_values().iloc[:min_len].reset_index(drop=True)
                    b_series = b_clean.sort_values().iloc[:min_len].reset_index(drop=True)
                    try:
                        if s_series.std() == 0 or b_series.std() == 0:
                            relational_stats["covariance"] = 0
                            relational_stats["correlation"] = 0
                        else:
                            cov = s_series.cov(b_series)
                            corr = s_series.corr(b_series)
                            relational_stats["covariance"] = 0 if pd.isna(cov) else cov
                            relational_stats["correlation"] = 0 if pd.isna(corr) else corr
                    except Exception:
                        pass

            p_opportunity = 0
            binom_results = {}
            if not sell_prices.empty:
                threshold = sell_prices.median() * 0.97
                success_count = len(sell_prices[sell_prices <= threshold])
                p_opportunity = success_count / len(sell_prices)

                if p_opportunity > 0:
                    p0 = (1 - p_opportunity) ** 10
                    binom_results['at_least_1'] = 1 - p0
                    binom_results['at_least_3'] = 1 - _binom_cdf(2, 10, p_opportunity)

            sell_metrics = self.calculate_metrics(sell_prices)
            buy_metrics = self.calculate_metrics(buy_prices)

            if sell_metrics: sell_metrics['mode'] = sell_mode
            if buy_metrics: buy_metrics['mode'] = buy_mode

            stats_dict = {
                "sell": sell_metrics,
                "buy": buy_metrics,
                "relationship": relational_stats,
                "p_opportunity": p_opportunity,
                "binom": binom_results,
                "sell_raw": sell_prices,
                "buy_raw": buy_prices
            }
            return stats_dict
        except Exception as e:
            print(f"Analiz hatası: {e}")
            return None

    def calculate_filtered_series(self, series):
        if series.empty: return series
        Q1, Q3 = series.quantile(0.25), series.quantile(0.75)
        IQR = Q3 - Q1
        lower = Q1 - self.iqr_multiplier * IQR
        upper = Q3 + self.iqr_multiplier * IQR
        filtered = series[(series >= lower) & (series <= upper)]
        return filtered if not filtered.empty else series

    def calculate_metrics(self, series):
        if series.empty:
            return None

        Q1, Q3 = series.quantile(0.25), series.quantile(0.75)
        IQR = Q3 - Q1
        lower_bound = Q1 - self.iqr_multiplier * IQR
        upper_bound = Q3 + self.iqr_multiplier * IQR

        filtered_series = series[(series >= lower_bound) & (series <= upper_bound)]

        if filtered_series.empty:
            filtered_series = series

        Q1 = filtered_series.quantile(0.25)
        Q3 = filtered_series.quantile(0.75)
        n = len(filtered_series)
        mean = filtered_series.mean()
        std_dev = filtered_series.std() if n > 1 else 0
        std_err = _sem(filtered_series) if n > 1 else 0

        ci_low, ci_high = (0, 0)
        if n > 1:
            ci_low, ci_high = _t_interval(0.95, n - 1, loc=mean, scale=std_err)

        return {
            "mean": mean,
            "median": filtered_series.median(),
            "mode": (filtered_series.mode().iloc[0] if not filtered_series.mode().empty else mean),
            "min": filtered_series.min(),
            "max": filtered_series.max(),
            "count": n,
            "q1": Q1,
            "q3": Q3,
            "variance": filtered_series.var() if n > 1 else 0,
            "std_dev": std_dev,
            "skew": filtered_series.skew() if n > 2 else 0,
            "kurt": filtered_series.kurt() if n > 2 else 0,
            "std_err": std_err,
            "ci_low": ci_low,
            "ci_high": ci_high,
            "dist_type": self.get_distribution_type(filtered_series)
        }

    def predict_deal(self, stats_data, user_offer=None):
        if not stats_data or not stats_data.get('sell') or not stats_data.get('buy'):
            return None

        max_buy = stats_data['buy']['max']
        sell_target = stats_data['sell']['median']
        suggested_offer = max_buy * 0.97

        target = user_offer if user_offer else suggested_offer
        sell_series = stats_data.get('sell_raw', pd.Series())

        p_success = 0
        if not sell_series.empty:
            success_count = len(sell_series[sell_series <= target])
            p_success = success_count / len(sell_series)

        return {
            "önerilen_alış": suggested_offer,
            "önerilen_satış": sell_target,
            "tahmini_kar": sell_target - suggested_offer,
            "p_opportunity": p_success,
            "bernoulli_variance": p_success * (1 - p_success)
        }

    def calculate_binom_prob(self, n, p, k):
        return 1 - _binom_cdf(k-1, n, p)

    def calculate_advanced_stats(self, buy_prices, sell_prices):
        stats_data = {}
        if not sell_prices.empty:
            stats_data['std_err'] = sell_prices.std()
            stats_data['variance'] = sell_prices.var()

        min_sell = sell_prices.min() if not sell_prices.empty else 0
        max_buy = buy_prices.max() if not buy_prices.empty else 0

        net_sell = min_sell * 0.97
        profit = net_sell - max_buy
        margin = (profit / max_buy * 100) if max_buy > 0 else 0

        stats_data['net_sell_price'] = net_sell
        stats_data['profit_per_unit'] = profit
        stats_data['margin_percent'] = margin
        return stats_data

    def get_all_unique_items(self):
        try:
            with self.db.get_connection() as conn:
                cursor = conn.cursor()
                cursor.execute("SELECT DISTINCT item_name, item_lvl FROM prices")
                rows = cursor.fetchall()
                return [{"name": row[0].strip() if row[0] else row[0], "lvl": row[1].strip() if row[1] else row[1]} for row in rows]
        except:
            return []

    def apply_manipulation_wall(self, series, wall_rate=0.01):
        if series.empty or len(series) < 10:
            return series
        lower_limit = series.quantile(wall_rate / 2)
        upper_limit = series.quantile(1 - (wall_rate / 2))
        filtered_series = series[(series >= lower_limit) & (series <= upper_limit)]
        return filtered_series

    def get_distribution_type(self, series):
        if series.empty or len(series) < 3: return "Yetersiz Veri"
        skewness = series.skew()
        if abs(skewness) < 0.5:
            return "Normal Dağılım (Dengeli Pazar)"
        elif skewness > 0.5:
            return "Pozitif Çarpık (Ucuz İlanlar Yoğunlukta)"
        else:
            return "Negatif Çarpık (Pahalı İlanlar Yoğunlukta)"

    def calculate_opportunity_score(self, stats, buy_price):
        try:
            if not stats or not stats.get("sell"):
                return {"score": 0, "rating": "Veri Yetersiz", "cv": 0}

            sell = stats["sell"]
            median_sell = sell.get("median", 0)
            expected_profit = median_sell - buy_price
            margin = (expected_profit / buy_price) * 100 if buy_price > 0 else 0
            profit_score = min(max(margin * 2, 0), 40)

            liquidity = sell.get("count", 0)
            liquidity_score = min(liquidity, 20)

            mean = sell.get("mean", 0)
            cv = (sell.get("std_dev", 0) / mean if mean > 0 else 1)
            stability_score = max(0, 20 - (cv * 100))

            p = stats.get("p_opportunity", 0)
            probability_score = p * 20

            total_score = round(min(max(profit_score + liquidity_score + stability_score + probability_score, 0), 100), 1)

            if total_score >= 80: rating = "🔥 ELITE FIRSAT"
            elif total_score >= 65: rating = "🟢 Güçlü Fırsat"
            elif total_score >= 45: rating = "🟡 Stabil"
            elif total_score >= 25: rating = "🟠 Riskli"
            else: rating = "🔴 Uzak Dur"

            return {"score": total_score, "rating": rating, "margin": margin, "cv": cv}
        except Exception as e:
            return {"score": 0, "rating": f"Hata: {e}", "cv": 0}

    def export_items_to_excel(self, selected_items, output_file="market_analysis.xlsx", server=None):
        export_rows = []

        for item in selected_items:
            item_name = item["name"]
            item_lvl = item["lvl"]

            stats_data = self.get_item_stats(item_name, item_lvl, server=server)

            if not stats_data:
                continue

            buy = stats_data.get("buy") or {}
            sell = stats_data.get("sell") or {}

            row = {
                "Tarih": pd.Timestamp.now().strftime("%Y-%m-%d %H:%M"),
                "Item": item_name,
                "Level": item_lvl,
                "BUY Min": buy.get("min", 0),
                "SELL Min": sell.get("min", 0),
                "BUY Max": buy.get("max", 0),
                "SELL Max": sell.get("max", 0),
                "BUY Mean": buy.get("mean", 0),
                "SELL Mean": sell.get("mean", 0),
                "BUY Median": buy.get("median", 0),
                "SELL Median": sell.get("median", 0),
                "BUY Mode": buy.get("mode", 0),
                "SELL Mode": sell.get("mode", 0),
                "BUY Q1": buy.get("q1", 0),
                "SELL Q1": sell.get("q1", 0),
                "BUY Q3": buy.get("q3", 0),
                "SELL Q3": sell.get("q3", 0),
                "BUY CI Low": buy.get("ci_low", 0),
                "SELL CI Low": sell.get("ci_low", 0),
                "BUY CI High": buy.get("ci_high", 0),
                "SELL CI High": sell.get("ci_high", 0),
                "BUY Variance": buy.get("variance", 0),
                "SELL Variance": sell.get("variance", 0),
                "BUY Std Dev": buy.get("std_dev", 0),
                "SELL Std Dev": sell.get("std_dev", 0),
                "Opportunity %": round(stats_data.get("p_opportunity", 0) * 100, 2),
                "Correlation": stats_data.get("relationship", {}).get("correlation", 0),
                "Covariance": stats_data.get("relationship", {}).get("covariance", 0)
            }

            export_rows.append(row)

        if not export_rows:
            return False

        df = pd.DataFrame(export_rows)

        base_name, ext = os.path.splitext(output_file)
        if ext == '': ext = '.xlsx'
        tarih_str = datetime.now().strftime("%d_%m_%Y")
        final_output_file = f"{base_name}_{tarih_str}{ext}"

        with pd.ExcelWriter(final_output_file, engine="openpyxl") as writer:
            df.to_excel(
                writer,
                sheet_name="Market Analysis",
                index=False
            )

            worksheet = writer.sheets["Market Analysis"]

            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                for cell in row:
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '#,##0'

            for column_cells in worksheet.columns:
                length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
                worksheet.column_dimensions[column_cells[0].column_letter].width = length + 5

        return final_output_file
